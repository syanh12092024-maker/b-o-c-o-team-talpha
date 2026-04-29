"""
Protokol Parser — Parse euShipments COD Payment Protocol files (XLSX + PDF fallback).

Handles XLSX (Bulgarian + English variants) and PDF (when XLSX is missing).
Extracts protocol_number & client_id from filename pattern: Protokol-{num}-{clientid}.{ext}

Usage:
    from protokol_parser import parse_protokol
    result = parse_protokol("Protokol-43472-3282.xlsx")   # XLSX
    result = parse_protokol("Protokol-43472-3282.pdf")    # PDF fallback
"""

import re
import os
import sys
import json
import logging
from pathlib import Path

# Windows UTF-8
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

logger = logging.getLogger(__name__)

# ─── Column indices (0-based) ─────────────────────────────────────────────────
COL_ROW_NUM = 0   # A: Row number
COL_AWB = 1       # B: Waybill number
COL_RECEIVER = 2  # C: Receiver/Recipient name
COL_NOTES = 5     # F: Notes / Order Info (contains POS order ID)
COL_AMOUNT = 6    # G: COD value / Amount (RON)
COL_CURRENCY = 7  # H: Currency
COL_COUNTED = 8   # I: Checked COD value / Counted


def extract_from_filename(filename: str) -> dict:
    """Extract protocol_number and client_id from filename.

    Pattern: Protokol-{protocol_number}-{client_id}.xlsx
    Example: Protokol-43472-3282.xlsx -> {"protocol_number": "43472", "client_id": "3282"}
    """
    basename = Path(filename).stem  # Remove extension
    m = re.search(r"Protokol-(\d+)-(\d+)", basename, re.IGNORECASE)
    if m:
        return {"protocol_number": m.group(1), "client_id": m.group(2)}
    # Fallback: try just protocol number
    m = re.search(r"Protokol-(\d+)", basename, re.IGNORECASE)
    if m:
        return {"protocol_number": m.group(1), "client_id": None}
    return {"protocol_number": None, "client_id": None}


def extract_pos_order_id(notes: str) -> str | None:
    """Extract POS order ID from notes column.

    Pattern: "SI_LEVELUPVN Ltd - Aurelia Wear (aureliastore) / 4166"
    Returns: "4166"
    """
    if not notes:
        return None
    m = re.search(r"/\s*(\d+)\s*$", str(notes).strip())
    return m.group(1) if m else None


def detect_variant(ws) -> str:
    """Detect XLSX variant by inspecting first rows.

    Returns 'english' or 'bulgarian'.
    """
    cell_a1 = ws.cell(row=1, column=1).value
    if cell_a1 is not None:
        s = str(cell_a1).strip().lower()
        if s.startswith("id") or "waybill" in s or "#" in s:
            return "english"
    # Check if Row 8 has header-like content (Bulgarian)
    cell_a8 = ws.cell(row=8, column=1).value
    if cell_a8 is not None:
        return "bulgarian"
    # Default: try English
    return "english"


def safe_float(val) -> float | None:
    """Convert value to float, handling commas and None."""
    if val is None:
        return None
    try:
        return float(str(val).replace(",", ".").strip())
    except (ValueError, TypeError):
        return None


def extract_eur_from_pdf(xlsx_path: str, xlsx_filename: str) -> tuple:
    """Try to extract paid_eur and exchange_rate from companion PDF.

    Looks for PDF with same name pattern in same directory.
    PDF footer contains: "Paid\\n4693.47 EUR\\nExchange rate\\nRate for RON: 0.19330"
    """
    try:
        import pdfplumber
    except ImportError:
        return None, None

    # Derive PDF path from XLSX path (same name, .pdf extension, remove -en suffix)
    pdf_name = re.sub(r"-en\.xlsx$", ".pdf", xlsx_filename, flags=re.IGNORECASE)
    pdf_name = re.sub(r"\.xlsx$", ".pdf", pdf_name, flags=re.IGNORECASE)
    pdf_path = os.path.join(os.path.dirname(xlsx_path), pdf_name)

    if not os.path.exists(pdf_path):
        return None, None

    try:
        pdf = pdfplumber.open(pdf_path)
        # Check last pages for summary
        for page in reversed(pdf.pages):
            text = page.extract_text() or ""
            lines = text.strip().split("\n")

            paid_eur = None
            exchange_rate = None

            for i, line in enumerate(lines):
                # "4693.47 EUR"
                m = re.match(r"^([\d.,]+)\s*EUR\s*$", line.strip())
                if m:
                    paid_eur = float(m.group(1).replace(",", ""))

                # "Rate for RON: 0.19330"
                m = re.search(r"Rate for RON:\s*([\d.]+)", line)
                if m:
                    exchange_rate = float(m.group(1))

            if paid_eur is not None:
                pdf.close()
                return paid_eur, exchange_rate

        pdf.close()
    except Exception as e:
        logger.debug(f"PDF parse failed for {pdf_path}: {e}")

    return None, None


def parse_protokol_xlsx(filepath: str, filename: str = None) -> dict:
    """Parse a Protokol XLSX file.

    Args:
        filepath: Path to the XLSX file on disk.
        filename: Original filename (for extracting protocol_number/client_id).
                  If None, uses basename of filepath.

    Returns:
        dict with keys:
            protocol_number, client_id, protocol_date,
            items: [{awb, receiver, amount_ron, counted_ron, currency, pos_order_id}],
            totals: {awb_count, total_amount_ron, total_counted_ron}
    """
    from openpyxl import load_workbook

    if filename is None:
        filename = os.path.basename(filepath)

    # Extract metadata from filename
    meta = extract_from_filename(filename)
    protocol_number = meta["protocol_number"]
    client_id = meta["client_id"]

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    variant = detect_variant(ws)
    logger.info(f"Detected variant: {variant} for {filename}")

    # Determine data start row
    if variant == "bulgarian":
        data_start_row = 9  # Rows 1-7 metadata, Row 8 headers
        # Try to extract protocol_date from Row 3
        protocol_date = None
        for row_idx in range(1, 8):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val:
                m = re.search(r"(\d{2}\.\d{2}\.\d{4})", str(cell_val))
                if m:
                    from datetime import datetime
                    try:
                        protocol_date = datetime.strptime(m.group(1), "%d.%m.%Y").strftime("%Y-%m-%d")
                    except ValueError:
                        pass
                    break
    else:
        data_start_row = 2  # Row 1 is header
        protocol_date = None

    # Parse data rows
    items = []
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        # Skip empty rows
        if not row or len(row) < COL_COUNTED + 1:
            continue

        awb_val = row[COL_AWB] if len(row) > COL_AWB else None
        if awb_val is None:
            continue

        awb = str(awb_val).strip()
        if not awb or not awb[0].isdigit():
            continue  # Skip non-AWB rows (like totals/footers)

        receiver = str(row[COL_RECEIVER] or "").strip()
        notes = str(row[COL_NOTES] or "").strip() if len(row) > COL_NOTES else ""
        amount = safe_float(row[COL_AMOUNT] if len(row) > COL_AMOUNT else None)
        currency = str(row[COL_CURRENCY] or "RON").strip() if len(row) > COL_CURRENCY else "RON"
        counted = safe_float(row[COL_COUNTED] if len(row) > COL_COUNTED else None)
        pos_order_id = extract_pos_order_id(notes)

        if amount is None:
            continue  # Skip rows without amount

        items.append({
            "awb": awb,
            "receiver": receiver,
            "amount_ron": amount,
            "counted_ron": counted,
            "currency": currency,
            "pos_order_id": pos_order_id,
            "notes": notes,
        })

    wb.close()

    # Compute totals
    total_amount = sum(i["amount_ron"] for i in items if i["amount_ron"])
    total_counted = sum(i["counted_ron"] for i in items if i["counted_ron"])

    # Try to extract paid_eur and exchange_rate from companion PDF
    paid_eur, exchange_rate = extract_eur_from_pdf(filepath, filename)

    result = {
        "protocol_number": protocol_number,
        "client_id": client_id,
        "protocol_date": protocol_date,
        "paid_eur": paid_eur,
        "exchange_rate": exchange_rate,
        "items": items,
        "totals": {
            "awb_count": len(items),
            "total_amount_ron": round(total_amount, 2),
            "total_counted_ron": round(total_counted, 2),
        },
    }

    logger.info(
        f"Parsed {filename}: protocol={protocol_number}, client={client_id}, "
        f"items={len(items)}, total_ron={total_amount:.2f}"
    )

    return result


# ─── PDF Parser (fallback when XLSX missing) ─────────────────────────────────

def parse_protokol_pdf(filepath: str, filename: str = None) -> dict:
    """Parse a Protokol PDF file using pdfplumber table extraction.

    Fallback for when XLSX is not available (e.g., old account 3248).
    PDF has tables with columns: #, Waybill number, Receiver, Amount, Currency, Counted
    Footer has: Total liability (RON), Paid (EUR), Exchange rate.
    """
    import pdfplumber

    if filename is None:
        filename = os.path.basename(filepath)

    meta = extract_from_filename(filename)
    protocol_number = meta["protocol_number"]
    client_id = meta["client_id"]

    pdf = pdfplumber.open(filepath)
    items = []
    protocol_date = None
    paid_eur = None
    exchange_rate = None

    for page in pdf.pages:
        # Extract tables
        tables = page.extract_tables()
        for table in tables:
            for row in table:
                if not row or len(row) < 6:
                    continue

                # Skip header row
                if row[0] and str(row[0]).strip().lower() in ("#", "№", "id"):
                    continue

                awb = str(row[1] or "").strip()
                if not awb or not awb[0].isdigit():
                    continue

                receiver = str(row[2] or "").strip()
                amount = safe_float(row[3])
                currency = str(row[4] or "RON").strip()
                counted = safe_float(row[5])

                if amount is None:
                    continue

                items.append({
                    "awb": awb,
                    "receiver": receiver,
                    "amount_ron": amount,
                    "counted_ron": counted,
                    "currency": currency,
                    "pos_order_id": None,  # PDF doesn't have Notes column
                    "notes": "",
                })

        # Extract text for metadata (date, EUR, rate)
        text = page.extract_text() or ""
        lines = text.strip().split("\n")

        # Protocol date from header
        if not protocol_date:
            for line in lines[:5]:
                m = re.search(r"(\d{2}\.\d{2}\.\d{4})", line)
                if m:
                    from datetime import datetime as dt
                    try:
                        protocol_date = dt.strptime(m.group(1), "%d.%m.%Y").strftime("%Y-%m-%d")
                    except ValueError:
                        pass

        # EUR and exchange rate from footer
        for line in lines:
            m = re.match(r"^([\d.,]+)\s*EUR\s*$", line.strip())
            if m:
                paid_eur = float(m.group(1).replace(",", ""))
            m = re.search(r"Rate for RON:\s*([\d.]+)", line)
            if m:
                exchange_rate = float(m.group(1))

    pdf.close()

    total_amount = sum(i["amount_ron"] for i in items if i["amount_ron"])
    total_counted = sum(i["counted_ron"] for i in items if i["counted_ron"])

    result = {
        "protocol_number": protocol_number,
        "client_id": client_id,
        "protocol_date": protocol_date,
        "paid_eur": paid_eur,
        "exchange_rate": exchange_rate,
        "items": items,
        "totals": {
            "awb_count": len(items),
            "total_amount_ron": round(total_amount, 2),
            "total_counted_ron": round(total_counted, 2),
        },
    }

    logger.info(
        f"Parsed PDF {filename}: protocol={protocol_number}, client={client_id}, "
        f"items={len(items)}, total_ron={total_amount:.2f}, paid_eur={paid_eur}"
    )

    return result


# ─── Compensation PDF Parser ──────────────────────────────────────────────────

def parse_compensation_pdf(filepath: str, filename: str = None) -> dict:
    """Parse a Compensation Protocol PDF.

    Contains: COD amount, list of invoices deducted, net amount paid.
    Format:
        COD Payment Protocol   Date        Amount
        43499                  18.3.2026   3,606.40 EUR
        Invoice     Date       Amount
        2000018184  9.3.2026   327.62 EUR
        Amount paid after deduction   479.84 EUR
    """
    import pdfplumber

    if filename is None:
        filename = os.path.basename(filepath)

    pdf = pdfplumber.open(filepath)
    full_text = ""
    for page in pdf.pages:
        full_text += (page.extract_text() or "") + "\n"
    pdf.close()

    lines = full_text.strip().split("\n")

    protocol_number = None
    protocol_date = None
    receiver = None
    cod_gross_eur = None
    invoices = []
    total_invoices_eur = None
    net_paid_eur = None

    for line in lines:
        stripped = line.strip()

        # Protocol number: "Protocol #: 43367" or "Protokol number: 43367"
        m = re.search(r"Protocol\s*#:\s*(\d+)", stripped, re.IGNORECASE)
        if m and not protocol_number:
            protocol_number = m.group(1)
        m = re.search(r"Protokol\s+number:\s*(\d+)", stripped, re.IGNORECASE)
        if m and not protocol_number:
            protocol_number = m.group(1)

        # Protocol date
        m = re.search(r"Date:\s*([\d.]+\.\d{2,4})", stripped)
        if m and not protocol_date:
            protocol_date = m.group(1)

        # Receiver company name
        if ("PTE" in stripped or "LIMITED" in stripped or "LTD" in stripped) and not receiver:
            receiver = stripped

        # COD total: "Total cash on delivery amount 1014.66 EUR"
        if "total" in stripped.lower() and "cash on delivery" in stripped.lower():
            m = re.search(r"([\d\s,.]+)\s*EUR", stripped)
            if m:
                cod_gross_eur = float(m.group(1).replace(" ", "").replace(",", ""))

        # COD protocol line: "43367 12.03.2026 1014.66 EUR"
        if not cod_gross_eur:
            m = re.match(r"^(\d{5})\s+[\d.]+\.\d{2,4}\s+([\d\s,.]+)\s*EUR", stripped)
            if m:
                cod_gross_eur = float(m.group(2).replace(" ", "").replace(",", ""))

        # Invoice lines: "INV 2000018266/2026-03-10 362.54 EUR 10.03.2026 362.54 EUR"
        inv_match = re.match(r"^INV\s+(\d+)(?:/[\d-]+)?\s+([\d,.]+)\s*EUR", stripped)
        if not inv_match:
            # Also try: "2000018184 9.3.2026 327.62 EUR"
            inv_match = re.match(r"^(\d{10,})\s+[\d.]+\.\d{4}\s+([\d\s,.]+)\s*EUR", stripped)
        if inv_match:
            invoices.append({
                "invoice_number": inv_match.group(1),
                "invoice_date": "",
                "amount_eur": float(inv_match.group(2).replace(" ", "").replace(",", "")),
            })

        # Total invoices: "Total offset amount 362.54 EUR" or "Total amount of invoices withheld"
        if "total" in stripped.lower() and ("offset" in stripped.lower() or "invoices withheld" in stripped.lower()):
            m = re.search(r"([\d\s,.]+)\s*EUR", stripped)
            if m:
                total_invoices_eur = float(m.group(1).replace(" ", "").replace(",", ""))

        # Net paid: "Amount Paid After Deduction 652.12 EUR"
        if "after deduction" in stripped.lower():
            m = re.search(r"([\d\s,.]+)\s*EUR", stripped)
            if m:
                net_paid_eur = float(m.group(1).replace(" ", "").replace(",", ""))

    # Fallback: compute total invoices if not found
    if total_invoices_eur is None and invoices:
        total_invoices_eur = round(sum(inv["amount_eur"] for inv in invoices), 2)

    result = {
        "type": "compensation",
        "protocol_number": protocol_number,
        "receiver": receiver,
        "protocol_date": protocol_date,
        "cod_gross_eur": cod_gross_eur,
        "invoices_deducted": invoices,
        "total_deductions_eur": total_invoices_eur,
        "net_paid_eur": net_paid_eur,
    }

    logger.info(
        f"Parsed compensation {filename}: protocol={protocol_number}, "
        f"COD={cod_gross_eur} EUR, deductions={total_invoices_eur} EUR, "
        f"net={net_paid_eur} EUR, invoices={len(invoices)}"
    )

    return result


# ─── Invoice PDF Parser ──────────────────────────────────────────────────────

def parse_invoice_pdf(filepath: str, filename: str = None) -> dict:
    """Parse an Invoice PDF from euShipments (InOut Trade LTD).

    Extracts invoice number, date, bill_to, line items, and totals.
    """
    import pdfplumber

    if filename is None:
        filename = os.path.basename(filepath)

    pdf = pdfplumber.open(filepath)
    full_text = ""
    tables_all = []
    for page in pdf.pages:
        full_text += (page.extract_text() or "") + "\n"
        tables_all.extend(page.extract_tables() or [])
    pdf.close()

    lines = full_text.strip().split("\n")

    invoice_number = None
    invoice_date = None
    bill_to = None
    period = None
    vat_base = None
    vat_amount = None
    balance_due = None
    line_items = []

    for line in lines:
        stripped = line.strip()

        # Invoice number: "INVOICE ... No: 2000018185"
        if "invoice" in stripped.lower() or (not invoice_number and re.search(r"No:\s*\d{10}", stripped)):
            m = re.search(r"No:\s*(\d{10,})", stripped)
            if m:
                invoice_number = m.group(1)

        # Invoice date
        m = re.search(r"Date:\s*(\d{2}\.\d{2}\.\d{4})", stripped)
        if m and not invoice_date:
            invoice_date = m.group(1)

        # Period from description
        if not period:
            m = re.search(r"for\s+((?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4})", stripped, re.IGNORECASE)
            if m:
                period = m.group(1)

        # VAT base
        m = re.search(r"VAT\s+base:\s*([\d,.]+)\s*EUR", stripped)
        if m:
            vat_base = float(m.group(1).replace(",", ""))

        # VAT amount
        m = re.search(r"VAT:\s*([\d,.]+)\s*EUR", stripped)
        if m:
            vat_amount = float(m.group(1).replace(",", ""))

        # Balance due
        m = re.search(r"Balance\s+due:\s*([\d,.]+)\s*EUR", stripped)
        if m:
            balance_due = float(m.group(1).replace(",", ""))

    # Extract bill_to from "BILL TO" section
    for i, line in enumerate(lines):
        if "BILL TO" in line:
            # Next non-empty line is company name
            for j in range(i + 1, min(i + 5, len(lines))):
                candidate = lines[j].strip()
                if candidate and not candidate.startswith("ISSUED") and len(candidate) > 3:
                    bill_to = candidate.split("Inout")[0].split("InOut")[0].strip()
                    break
            break

    # Extract line items from tables
    for table in tables_all:
        for row in table:
            if not row or len(row) < 5:
                continue
            # Skip header rows
            if row[0] and str(row[0]).strip().lower() in ("#", "item", "№"):
                continue
            # Try to parse: #, Description, Qty, Unit Cost, Price
            try:
                row_num = str(row[0] or "").strip()
                if not row_num or not row_num[0].isdigit():
                    continue
                desc = str(row[1] or "").strip()
                if not desc:
                    continue
                # Find numeric columns from the right
                nums = []
                for cell in reversed(row[2:]):
                    if cell is not None:
                        try:
                            nums.insert(0, float(str(cell).replace(",", "").strip()))
                        except (ValueError, TypeError):
                            break
                if len(nums) >= 3:
                    line_items.append({
                        "description": desc,
                        "qty": nums[-3],
                        "unit_cost": nums[-2],
                        "total": nums[-1],
                    })
                elif len(nums) >= 1:
                    line_items.append({
                        "description": desc,
                        "qty": None,
                        "unit_cost": None,
                        "total": nums[-1],
                    })
            except (IndexError, ValueError):
                continue

    # Fallback: extract invoice number from filename
    if not invoice_number:
        m = re.search(r"inv-(\d+)", filename, re.IGNORECASE)
        if m:
            invoice_number = m.group(1)

    result = {
        "type": "invoice",
        "invoice_number": invoice_number,
        "invoice_date": invoice_date,
        "bill_to": bill_to,
        "period": period,
        "line_items": line_items,
        "vat_base_eur": vat_base,
        "vat_eur": vat_amount,
        "total_eur": balance_due or vat_base,
        "source_filename": filename,
    }

    logger.info(
        f"Parsed invoice {filename}: #{invoice_number}, "
        f"bill_to={bill_to}, items={len(line_items)}, total={balance_due} EUR"
    )

    return result


# ─── Unified Parser ──────────────────────────────────────────────────────────

def parse_protokol(filepath: str, filename: str = None) -> dict:
    """Parse Protokol file — auto-detects XLSX or PDF."""
    if filename is None:
        filename = os.path.basename(filepath)

    if filename.lower().endswith(".pdf"):
        return parse_protokol_pdf(filepath, filename)
    else:
        return parse_protokol_xlsx(filepath, filename)


# ─── CLI ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse

    logging.basicConfig(level=logging.INFO, format="%(message)s")

    ap = argparse.ArgumentParser(description="Parse Protokol/Compensation/Invoice PDF/XLSX")
    ap.add_argument("--file", required=True, help="Path to file")
    ap.add_argument("--type", choices=["auto", "protokol", "compensation", "invoice"], default="auto",
                    help="Document type (default: auto-detect)")
    ap.add_argument("--json", action="store_true", help="Output as JSON")
    args = ap.parse_args()

    filepath = args.file
    filename = os.path.basename(filepath)

    if args.type == "compensation":
        result = parse_compensation_pdf(filepath, filename)
    elif args.type == "invoice":
        result = parse_invoice_pdf(filepath, filename)
    elif args.type == "auto" and filename.lower().startswith("inv-"):
        result = parse_invoice_pdf(filepath, filename)
    elif args.type == "auto" and "compensation" in filename.lower():
        result = parse_compensation_pdf(filepath, filename)
    else:
        result = parse_protokol(filepath, filename)

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        doc_type = result.get("type", "protokol")
        if doc_type == "compensation":
            print(f"COMPENSATION Protocol #{result['protocol_number']}")
            print(f"  COD Gross:    EUR {result['cod_gross_eur']}")
            print(f"  Deductions:   EUR {result['total_deductions_eur']} ({len(result['invoices_deducted'])} invoices)")
            for inv in result["invoices_deducted"]:
                print(f"    - INV {inv['invoice_number']}: EUR {inv['amount_eur']}")
            print(f"  Net Paid:     EUR {result['net_paid_eur']}")
        elif doc_type == "invoice":
            print(f"INVOICE #{result['invoice_number']} | {result['invoice_date']}")
            print(f"  Bill To: {result['bill_to']}")
            print(f"  Period:  {result['period']}")
            print(f"  Total:   EUR {result['total_eur']}")
            print(f"  Items:")
            for item in result["line_items"]:
                print(f"    - {item['description']}: EUR {item['total']}")
        else:
            print(f"PROTOKOL #{result.get('protocol_number')} | Client: {result.get('client_id')}")
            print(f"  Items: {result['totals']['awb_count']} | RON {result['totals']['total_amount_ron']:,.2f}")
            if result.get("paid_eur"):
                print(f"  Paid: EUR {result['paid_eur']}")
